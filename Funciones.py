import random    #Importamos el paquete random para obtener un número aleatorio
import pandas as pd    #Importamos el paquete pandas para manejar dataframes
import openpyxl
import os    #Importamos el paquete os para trabajar con rutas
from getpass import getpass     #Importamos del paquete getpass, la función getpass para así hacer que en el juego en parejas, no se muestren por pantalla los intentos
from openpyxl import load_workbook    #Importampos del paquete openpyxl la función load_workbook para cargar un excel.
from tabulate import tabulate   # Para enseñar un df por pantalla.

#Definiremos una función que pida al usuario una cadena de strings que será definido como nombre. La función nos devuelve la variable nombre.
def pedir_nombre_solitario():

    nombre = input("Hola! Cómo te llamas? ")
    return nombre
#Definiremos una función que pida al usuario dos cadenas de strings que serán definidos como nombre_1 y nombre_2. La función nos devuelve las variables nombre_1 y nombre_2.
def pedir_nombre_parejas():

    nombre_1 = input('¿Cómo se llama el primer integrante? ')
    nombre_2 = input('¿Cómo se llama el segundo integrante? ')
    return nombre_1,nombre_2
#Función para pedir el nivel de la partida al jugador en modo solitario. Tiene como parámetro nombre, así no se tiene que volver a preguntar.
#Enseña un menú a partir de varios prints, pide un valor de caracter integrer. Dependidendo de este valor se definen las variables
#dificultad y numero_intentos, ambas de tipo integrer, de manera diferente. La función decuelve las variables dificultad y numero_intentos.
def pedir_nivel_solitario(nombre):

    opcion_valida=False
    print(f'Hola {nombre}, ha llegado la hora de decidir la dificultad del juego.')
    while opcion_valida == False:
        print("\n--- Nivel de dificultad ---")
        print("1. Pulsa 1 si quieres jugar el modo fácil y tener 20 intentos.")
        print("2. Pulsa 2 si quieres jugar el modo intermedio y tener 12 intentos.")
        print("3. Pulsa 3 si quieres jugar el modo difícil y tener 5 intentos.")

        opcion = int(input("Elige una opción: "))

        if opcion == 1:
            dificultad = 1
            numero_intentos = 20
            print(f'Has elegido la dificultad fácil, tienes {numero_intentos} intentos')
            opcion_valida=True
        elif opcion == 2:
            dificultad = 2
            numero_intentos = 12
            print(f'Has elegido la dificultad intermedia, tienes {numero_intentos} intentos')
            opcion_valida=True
        elif opcion == 3:
            dificultad = 3
            numero_intentos = 5
            print(f'Has elegido la dificultad intermedia, tienes {numero_intentos} intentos')
            opcion_valida=True
        else:
            print("❌ Opción inválida, intenta de nuevo.")
    return numero_intentos, dificultad
#Función para pedir el nivel de la partida a los jugadores en modo pareja. Tiene como parámetros nombre_1 y nombre_2, así no se tiene que volver a preguntar.
#Enseña un menú a partir de varios prints, pide un valor de caracter integrer. Dependidendo de este valor se definen las variables
#dificultad y numero_intentos, ambas de tipo integrer, de manera diferente. La función decuelve las variables dificultad y numero_intentos
def pedir_nivel_pareja(nombre_1,nombre_2):

    opcion_valida=False
    print(f'Hola {nombre_1} y {nombre_2} , ha llegado la hora de decidir la dificultad del juego.')
    while opcion_valida == False:
        print("\n--- Nivel de dificultad ---")
        print("1. Pulsa 1 si queréis jugar el modo fácil y tener 20 intentos.")
        print("2. Pulsa 2 si queréis jugar el modo intermedio y tener 12 intentos.")
        print("3. Pulsa 3 si queréis jugar el modo difícil y tener 5 intentos.")

        opcion = int(input("Elige una opción: "))

        if opcion == 1:
            dificultad = 1
            numero_intentos = 20
            print(f'Habéis elegido la dificultad fácil, teneis {numero_intentos} intentos')
            opcion_valida=True
        elif opcion == 2:
            dificultad = 2
            numero_intentos = 12
            print(f'Hebéis elegido la dificultad intermedia, teneis {numero_intentos} intentos')
            opcion_valida=True
        elif opcion == 3:
            dificultad = 3
            numero_intentos = 5
            print(f'Habéis elegido la dificultad intermedia, tenéis {numero_intentos} intentos')
            opcion_valida=True
        else:
            print("❌ Opción inválida, intentadlo de nuevo.")
    return numero_intentos,dificultad
# Esta función sirve para, una vez acabe la partida en solitario, apunte el nombre del jugador en un excel con sus estadísticas. Si 
#este ya existe, simplemente se actualizan las estadísticas. Si este no existe, se le añade en el excel y se guardan las estadísticas de 
#la nueva partida. La función recibe como parametros nombre (string), dificultad (integrer) y encierto (booleano). A partir de esta información
#se añade el valor a una columna o a otra.
def resultado_jugadores_solitario(nombre,dificultad, encierto):
    archivo = "Resultados de adivina el número.xlsx"      #Aquí carga el excel, y abre la hoja "Resultado jugadores".
    wb = load_workbook(archivo)
    ws = wb["Resultado jugadores"]
    fila=0
    nombre_encontrado = False

    for celda in ws["A"]:                       #Aquí busca si el jugador ya existe.
        fila += 1 
        if celda.value == nombre:
            nombre_encontrado = True
            break
    if nombre_encontrado:               #Aquí, si el jugador se ha encontrado, actualiza las columnas
        ws.cell(row=fila, column=5).value += 1
        if dificultad == 1:
            ws.cell(row=fila, column=4).value += 1
        elif dificultad == 2:
            ws.cell(row=fila, column=3).value += 1        
        elif dificultad == 3:
            ws.cell(row=fila, column=2).value += 1
        if encierto:
            ws.cell(row=fila, column=7).value += 1
        num = ws.cell(row=fila, column=7).value 
        den = ws.cell(row=fila, column=5).value

        ws.cell(row=fila, column=9).value = num / den
    else:                               #Si no se ha encontrado, se añade una nueva columna
        if dificultad == 1 and encierto:
            nueva_fila = [nombre, 0, 0, 1, 1, 0, 1, 0, 1,0]
        if dificultad == 1 and not encierto:
            nueva_fila = [nombre, 0, 0, 1, 1, 0, 0, 0, 0,0]
        elif dificultad == 2 and encierto:
            nueva_fila = [nombre, 0, 1, 0, 1, 0, 1, 0, 1,0]
        elif dificultad == 2 and not encierto:
            nueva_fila = [nombre, 0, 1, 0, 1, 0, 0, 0, 0,0]
        elif dificultad == 3 and encierto:
            nueva_fila = [nombre, 1, 0, 0, 1, 0, 1, 0, 1,0]
        elif dificultad == 3 and not encierto:
            nueva_fila = [nombre, 1, 0, 0, 1, 0, 0, 0, 0,0]
        ws.append(nueva_fila)
    ws.cell(row=fila, column=9).number_format = "0.00%"
    ws.cell(row=fila, column=10).number_format = "0.00%"  
    wb.save(archivo)

#Esta función se encarga de crear un excel con el nombre "Resultados de adivina el número.xlsx" si este no existe en el entorno que trabajemos.
#Si el excel no existe, se crea uno con tres hojas: 'Solitario', 'Pareja' y 'Resultado jugadores'.

def crear_excel():     
    archivo = "Resultados de adivina el número.xlsx"

    if os.path.exists(archivo):
        return

    # Hoja: Solitario
    columnas_solitario = [
        "Nombre",
        "Dificultad",
        "Final",
        "Nº de intentos",
        "Numero secreto",
        "Último intento"
    ]

    # Hoja: Pareja
    columnas_pareja = [
        "Jugador 1",
        "Jugador 2",
        "Ganador/a",
        "Nº de intentos Jugador 1",
        "Nº de intentos Jugador 2",
        "Número Secreto",
        "Último intento"
    ]

    # Hoja: Resultado jugadores (estadísticas agregadas por jugador)
    columnas_resultado = [
        "Nombre",
        "Resultados partidas modo difícil",
        "Resultados partidas modo intermedio",
        "Resultados partidas modo fácil",
        "Partidas en solitario",
        "Partidas en pareja",
        "Victorias solitario",
        "Victorias en pareja",
        "Porcentaje de victorias solitario",
        "Porcentaje de victorias pareja"
    ]

    df_solitario = pd.DataFrame(columns=columnas_solitario)
    df_pareja = pd.DataFrame(columns=columnas_pareja)
    df_resultado = pd.DataFrame(columns=columnas_resultado)

    # Crear el Excel con las 3 hojas
    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        df_solitario.to_excel(writer, sheet_name="Solitario", index=False)
        df_pareja.to_excel(writer, sheet_name="Pareja", index=False)
        df_resultado.to_excel(writer, sheet_name="Resultado jugadores", index=False)
        
# Esta función sirve para, una vez acabe la partida en pareja, apunte el nombre de los jugadores en un excel con sus estadísticas. Si 
#estos ya existen, simplemente se actualizan las estadísticas. Si estos no existen, se les añade en el excel y se guardan las estadísticas de 
#la nueva partida. La función recibe como parametros nombre_1 y nombre_2 (string), dificultad (integrer), encierto_1 y enicerto_2 (booleano). A partir de esta información
#se añade el valor a una columna o a otra.
def resultado_jugadores_pareja(nombre_1,nombre_2,dificultad, encierto_1, encierto_2):
    archivo = "Resultados de adivina el número.xlsx"
    wb = load_workbook(archivo)
    ws = wb["Resultado jugadores"]

    fila_1 = 0
    fila_2 = 0
    nombre_encontrado_1 = False
    nombre_encontrado_2 = False

    # Buscar jugador 1
    for celda in ws["A"]:
        fila_1 += 1 
        if celda.value == nombre_1:
            nombre_encontrado_1 = True
            break
    # Buscar jugador 2
    for celda in ws["A"]:
        fila_2 += 1 
        if celda.value == nombre_2:
            nombre_encontrado_2 = True
            break

    # ---- Los dos existen ----
    if nombre_encontrado_1 and nombre_encontrado_2:
        for fila in [fila_1, fila_2]:
            # F: Partidas en pareja
            ws.cell(row=fila, column=6).value = (ws.cell(row=fila, column=6).value or 0) + 1
            # Dificultad (B,C,D)
            if dificultad == 1:
                ws.cell(row=fila, column=4).value = (ws.cell(row=fila, column=4).value or 0) + 1
            elif dificultad == 2:
                ws.cell(row=fila, column=3).value = (ws.cell(row=fila, column=3).value or 0) + 1
            elif dificultad == 3:
                ws.cell(row=fila, column=2).value = (ws.cell(row=fila, column=2).value or 0) + 1
            # H: Victorias pareja (segoun el jugador)
            if encierto_1 and fila == fila_1:
                ws.cell(row=fila, column=8).value = (ws.cell(row=fila, column=8).value or 0) + 1
            if encierto_2 and fila == fila_2:
                ws.cell(row=fila, column=8).value = (ws.cell(row=fila, column=8).value or 0) + 1
            # J = H/F
            num = ws.cell(row=fila, column=8).value or 0
            den = ws.cell(row=fila, column=6).value or 0
            ws.cell(row=fila, column=10).value = (num / den) if den else 0

    # ---- Solo uno existe ----
    elif (nombre_encontrado_1 and not nombre_encontrado_2) or (nombre_encontrado_2 and not nombre_encontrado_1):
        if nombre_encontrado_1:
            # actualiza jugador 1 existente
            ws.cell(row=fila_1, column=6).value = (ws.cell(row=fila_1, column=6).value or 0) + 1
            if dificultad == 1:
                ws.cell(row=fila_1, column=4).value = (ws.cell(row=fila_1, column=4).value or 0) + 1
            elif dificultad == 2:
                ws.cell(row=fila_1, column=3).value = (ws.cell(row=fila_1, column=3).value or 0) + 1
            elif dificultad == 3:
                ws.cell(row=fila_1, column=2).value = (ws.cell(row=fila_1, column=2).value or 0) + 1
            if encierto_1:
                ws.cell(row=fila_1, column=8).value = (ws.cell(row=fila_1, column=8).value or 0) + 1
            # J para jugador 1
            num = ws.cell(row=fila_1, column=8).value or 0
            den = ws.cell(row=fila_1, column=6).value or 0
            ws.cell(row=fila_1, column=10).value = (num / den) if den else 0
            ws.cell(row=fila_1, column=10).number_format = "0.00%"

            # crea fila para jugador 2 no existente (pareja: F=1, H según encierto_2)
            if dificultad == 1 and encierto_2:
                nueva_fila = [nombre_2, 0, 0, 1, 0, 1, 0, 1, 0, 1]
            elif dificultad == 1:
                nueva_fila = [nombre_2, 0, 0, 1, 0, 1, 0, 0, 0, 0]
            elif dificultad == 2 and encierto_2:
                nueva_fila = [nombre_2, 0, 1, 0, 0, 1, 0, 1, 0, 1]
            elif dificultad == 2:
                nueva_fila = [nombre_2, 0, 1, 0, 0, 1, 0, 0, 0, 0]
            elif dificultad == 3 and encierto_2:
                nueva_fila = [nombre_2, 1, 0, 0, 0, 1, 0, 1, 0, 1]
            else:
                nueva_fila = [nombre_2, 1, 0, 0, 0, 1, 0, 0, 0, 0]
            ws.append(nueva_fila)
            fila_2 = ws.max_row  # ← imprescindible

            # J para la fila nueva
            num = ws.cell(row=fila_2, column=8).value or 0
            den = ws.cell(row=fila_2, column=6).value or 0
            ws.cell(row=fila_2, column=10).value = (num / den) if den else 0
            ws.cell(row=fila_2, column=10).number_format = "0.00%"

        else:
            # actualiza jugador 2 existente  (FIX de sintaxis)
            ws.cell(row=fila_2, column=6).value = (ws.cell(row=fila_2, column=6).value or 0) + 1
            if dificultad == 1:
                ws.cell(row=fila_2, column=4).value = (ws.cell(row=fila_2, column=4).value or 0) + 1
            elif dificultad == 2:
                ws.cell(row=fila_2, column=3).value = (ws.cell(row[fila_2, 3]).value or 0) + 1  # <- también puedes usar column=3
            elif dificultad == 3:
                ws.cell(row=fila_2, column=2).value = (ws.cell(row=fila_2, column=2).value or 0) + 1
            if encierto_2:
                ws.cell(row=fila_2, column=8).value = (ws.cell(row=fila_2, column=8).value or 0) + 1
            # J para jugador 2
            num = ws.cell(row=fila_2, column=8).value or 0
            den = ws.cell(row=fila_2, column=6).value or 0
            ws.cell(row=fila_2, column=10).value = (num / den) if den else 0
            ws.cell(row=fila_2, column=10).number_format = "0.00%"

            # crea fila para jugador 1 NO existente (pareja: F=1, H según encierto_1)
            if dificultad == 1 and encierto_1:
                nueva_fila = [nombre_1, 0, 0, 1, 0, 1, 0, 1, 0, 1]
            elif dificultad == 1:
                nueva_fila = [nombre_1, 0, 0, 1, 0, 1, 0, 0, 0, 0]
            elif dificultad == 2 and encierto_1:
                nueva_fila = [nombre_1, 0, 1, 0, 0, 1, 0, 1, 0, 1]
            elif dificultad == 2:
                nueva_fila = [nombre_1, 0, 1, 0, 0, 1, 0, 0, 0, 0]
            elif dificultad == 3 and encierto_1:
                nueva_fila = [nombre_1, 1, 0, 0, 0, 1, 0, 1, 0, 1]
            else:
                nueva_fila = [nombre_1, 1, 0, 0, 0, 1, 0, 0, 0, 0]
            ws.append(nueva_fila)
            fila_1 = ws.max_row  # ← imprescindible

            # J para la fila nueva
            num = ws.cell(row=fila_1, column=8).value or 0
            den = ws.cell(row=fila_1, column=6).value or 0
            ws.cell(row=fila_1, column=10).value = (num / den) if den else 0
            ws.cell(row=fila_1, column=10).number_format = "0.00%"

    # ---- Cap existeix: crea dues files (⚠️ 1a per nombre_1, 2a per nombre_2) ----
    else:
        # jugador 1
        if dificultad == 1 and encierto_1:
            nueva_1 = [nombre_1, 0, 0, 1, 0, 1, 0, 1, 0, 1]
        elif dificultad == 1 and not encierto_1:
            nueva_1 = [nombre_1, 0, 0, 1, 0, 1, 0, 0, 0, 0]
        elif dificultad == 2 and encierto_1:
            nueva_1 = [nombre_1, 0, 1, 0, 0, 1, 0, 1, 0, 1]
        elif dificultad == 2 and not encierto_1:
            nueva_1 = [nombre_1, 0, 1, 0, 0, 1, 0, 0, 0, 0]
        elif dificultad == 3 and encierto_1:
            nueva_1 = [nombre_1, 1, 0, 0, 0, 1, 0, 1, 0, 1]
        else:
            nueva_1 = [nombre_1, 1, 0, 0, 0, 1, 0, 0, 0, 0]
        ws.append(nueva_1)
        fila_1 = ws.max_row

        # jugador 2
        if dificultad == 1 and encierto_2:
            nueva_2 = [nombre_2, 0, 0, 1, 0, 1, 0, 1, 0, 1]
        elif dificultad == 1 and not encierto_2:
            nueva_2 = [nombre_2, 0, 0, 1, 0, 1, 0, 0, 0, 0]
        elif dificultad == 2 and encierto_2:
            nueva_2 = [nombre_2, 0, 1, 0, 0, 1, 0, 1, 0, 1]
        elif dificultad == 2 and not encierto_2:
            nueva_2 = [nombre_2, 0, 1, 0, 0, 1, 0, 0, 0, 0]
        elif dificultad == 3 and encierto_2:
            nueva_2 = [nombre_2, 1, 0, 0, 0, 1, 0, 1, 0, 1]
        else:
            nueva_2 = [nombre_2, 1, 0, 0, 0, 1, 0, 0, 0, 0]
        ws.append(nueva_2)
        fila_2 = ws.max_row

    # Format percentatges (si les files són vàlides)
    if fila_1 > 0:
        ws.cell(row=fila_1, column=9).number_format = "0.00%"
        ws.cell(row=fila_1, column=10).number_format = "0.00%"
    if fila_2 > 0:
        ws.cell(row=fila_2, column=9).number_format = "0.00%"
        ws.cell(row=fila_2, column=10).number_format = "0.00%"

    wb.save(archivo)




def solitario():
    path='Resultados de adivina el número.xlsx'
    nombre = pedir_nombre_solitario()
    wb = load_workbook(path)
    niveles={1:'Fácil', 2:'Intermedio',3:'Difícil'}
    resultado={True:'Victoria', False:'Derrota'}
    numero_intentos,dificultad = pedir_nivel_solitario(nombre)
    print(f"\n🔢 El juego empieza con {numero_intentos} intentos.")
    encierto = False  #Esta variable nos indica que no se ha encontrado el num secreto. Si se encuentra, encierto pasara a ser True y se romperá el bucle siguiente.
    numero_elegido = random.randint(1,1001) #Se genera un número aleatorio entre el 1 y el 1000.
    i = 0  #Es el contador.
    while i in range(numero_intentos) and encierto == False:
        intento=int(input(f'Tu intento número {i+1} es:'))
        if intento == numero_elegido:
            os.system("afplay 'success-fanfare-trumpets-6185.mp3'")
            encierto=True
            print(f'Has acertado en tu intento {i+1}! El número secreto era {numero_elegido}!')
        if intento<numero_elegido:
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número {intento} es menor que el número secreto... Te quedan {numero_intentos-(i+1)} intentos!')
            i += 1
        if intento>numero_elegido:
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número {intento} es mayor que el número secreto... Te quedan {numero_intentos-(i+1)} intentos!')
            i += 1
    if encierto == False:
        print(f"Lo siento, no has adivinado el número secreto que era {numero_elegido}... Inténtalo otra vez!")
    ws=wb['Solitario']
    nueva_fila=[nombre,niveles[dificultad],resultado[encierto],i,numero_elegido,intento]
    ws.append(nueva_fila)
    wb.save(path)
    resultado_jugadores_solitario(nombre,dificultad, encierto)

def pareja():
    path='Resultados de adivina el número.xlsx'
    wb = load_workbook(path)
    niveles = {1:'Fácil', 2:'Intermedio',3:'Difícil'}
    resultado = {True:'Victoria', False:'Derrota'}
    nombre_1, nombre_2 = pedir_nombre_parejas()
    numero_intentos, dificultad = pedir_nivel_pareja(nombre_1, nombre_2)
    print(f"\n🔢 El juego empieza con {numero_intentos} intentos.")
    encierto_1 = False  #Esta variable nos indica que no se ha encontrado el num secreto. Si se encuentra, encierto pasara a ser True y se romperá el bucle siguiente.
    encierto_2 = False  
    numero_elegido = random.randint(1,1001) #Se genera un número aleatorio entre el 1 y el 1000.
    i = 0  #Es el contador del primer jugador.
    j=  0  #Es el contador del segundo jugador
    while j in range(numero_intentos) and (not encierto_1 and not encierto_2):
        intento_1=int(getpass(f'El intento número {i+ 1} de {nombre_1} es:'))
        if intento_1 == numero_elegido:
            i += 1
            os.system("afplay 'success-fanfare-trumpets-6185.mp3'")
            print(f'Enhorabuena {nombre_1}! Has acertado en tu intento {i}! El número secreto era {numero_elegido}!')
            ganador=nombre_1
            perdedor=nombre_2
            encierto_1 = True
            break
        if intento_1<numero_elegido:
            i += 1
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número que me has dado es menor que el número secreto... Te quedan {numero_intentos-(i)} intentos!')
        if intento_1>numero_elegido:
            i += 1
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número que me has dado es mayor que el número secreto... Te quedan {numero_intentos-(i)} intentos!')
        intento_2=int(getpass(f'El intento número {j+1} de {nombre_2} es:'))
        if intento_2 == numero_elegido:
            j += 1
            os.system("afplay 'success-fanfare-trumpets-6185.mp3'")
            print(f'Enhorabuena {nombre_2}! Has acertado en tu intento {j}! El número secreto era {numero_elegido}!')
            ganador = nombre_2
            perdedor = nombre_1
            encierto_2 = True
            break
        if intento_2<numero_elegido:
            j += 1
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número que me has dado es menor que el número secreto... Te quedan {numero_intentos-(j)} intentos!')
        if intento_2>numero_elegido:
            j += 1
            os.system("afplay 'mixkit-wrong-answer-bass-buzzer-948.wav'")
            print(f'El número que me has dado es mayor que el número secreto... Te quedan {numero_intentos-(j)} intentos!')
    if not encierto_1 and not encierto_2:
        print(f"Lo siento, no has adivinado el número secreto, que era {numero_elegido}... Inténtalo otra vez!")
        ganador = '-'
    ws=wb['Pareja']
    nueva_fila=[nombre_1,nombre_2,ganador,i,j,numero_elegido,f'{intento_1}-{intento_2}']
    ws.append(nueva_fila)
    wb.save(path)
    resultado_jugadores_pareja(nombre_1,nombre_2,dificultad, encierto_1, encierto_2)
def estadisticas():
    archivo = "Resultados de adivina el número.xlsx"

    dfs = pd.read_excel(archivo, sheet_name=["Solitario", "Pareja", "Resultado jugadores"])

    for nombre, df in dfs.items():
        print(f"\n========== {nombre} ==========")
        print(tabulate(df, headers='keys', tablefmt='psql'))      


 